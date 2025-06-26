import re
import requests
import random
import hashlib
import openpyxl
import time
import os
import sys
import textwrap
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import datetime

# ANSI escape codes for colored output
class Colors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'

def print_header(message):
    print(f"{Colors.OKBLUE}{'=' * 70}{Colors.ENDC}")
    print(f"{Colors.BOLD}{message.center(70)}{Colors.ENDC}")
    print(f"{Colors.OKBLUE}{'=' * 70}{Colors.ENDC}")

def print_success(message):
    print(f"{Colors.OKGREEN}✓ {message}{Colors.ENDC}")

def print_warning(message):
    print(f"{Colors.WARNING}⚠️ {message}{Colors.ENDC}")

def print_error(message):
    print(f"{Colors.FAIL}✗ {message}{Colors.ENDC}")

def baidu_translate(query, app_id, app_key):
    """使用百度翻译API进行中译英，带有频率控制和错误处理"""
    salt = random.randint(32768, 65536)
    sign_str = app_id + query + str(salt) + app_key
    sign = hashlib.md5(sign_str.encode()).hexdigest()
    
    url = "https://fanyi-api.baidu.com/api/trans/vip/translate"
    params = {
        'q': query,
        'from': 'zh',
        'to': 'en',
        'appid': app_id,
        'salt': salt,
        'sign': sign
    }
    
    try:
        response = requests.get(url, params=params, timeout=25)
        response.raise_for_status()
        result = response.json()
        
        if 'error_code' in result:
            error_code = result['error_code']
            error_msg = result.get('error_msg', 'Unknown error')
            
            if error_code == '54003':  # 访问频率受限
                print_error("错误54003: 访问频率受限")
                print_warning("原因: 您的请求频率超过了API限制")
                print_warning("解决方案:")
                print_warning("1. 增加请求间隔时间（免费用户1秒/次）")
                print_warning("2. 升级到企业版API")
                print_warning("3. 分批处理Excel文件")
                return "ERROR_54003"
            
            elif error_code == '52003':  # 未授权用户
                print_error("错误52003: 未授权用户")
                print_warning("可能原因:")
                print_warning("1. AppID或AppKey配置错误")
                print_warning("2. IP地址未添加到白名单")
                print_warning("3. 服务未开通或账户问题")
                return None
            
            else:
                print_error(f"翻译错误: {query[:30]}... | 错误代码: {error_code} | 错误信息: {error_msg}")
                return None
        
        return result['trans_result'][0]['dst']
    
    except requests.exceptions.RequestException as e:
        print_error(f"网络请求失败: {query[:30]}... | 错误: {str(e)}")
        return None
    except ValueError:
        print_error(f"JSON解析失败: {response.text[:100]}")
        return None
    except Exception as e:
        print_error(f"未知错误: {str(e)}")
        return None

def is_already_translated(cell_value):
    """检查单元格是否已经包含翻译内容"""
    if not isinstance(cell_value, str):
        return False
    
    # 检查是否包含英文括号格式：中文(英文)
    if re.search(r'\([a-zA-Z].*\)$', cell_value):
        return True
    
    # 检查是否包含换行分隔格式：英文\n中文（检查第一部分是否有英文）
    if '\n' in cell_value:
        parts = cell_value.split('\n')
        if len(parts) >= 2:
            first_part = parts[0].strip()
            if any(c.isalpha() for c in first_part):  # 检查第一部分是否包含英文字母
                return True
    
    return False

def check_and_adjust_translation_order(cell_value):
    """检查并调整中英文翻译顺序，确保英文在中文上方"""
    if not isinstance(cell_value, str):
        return cell_value
    
    # 处理括号格式：中文(英文)
    bracket_match = re.match(r'^(.*?)\s*\(([A-Za-z\s].*?)\)\s*$', cell_value)
    if bracket_match:
        chinese_part = bracket_match.group(1).strip()
        english_part = bracket_match.group(2).strip()
        return f"{english_part}\n{chinese_part}"
    
    # 处理换行格式
    if '\n' in cell_value:
        parts = [part.strip() for part in cell_value.split('\n') if part.strip()]
        if len(parts) == 2:
            part1, part2 = parts
            # 检查哪部分是中文，哪部分是英文
            has_chinese1 = bool(re.search(r'[\u4e00-\u9fa5]', part1))
            has_chinese2 = bool(re.search(r'[\u4e00-\u9fa5]', part2))
            
            # 如果part1是中文，part2是英文，则交换
            if has_chinese1 and not has_chinese2:
                return f"{part2}\n{part1}"
            # 如果已经是英文在上，中文在下，则不改变
            elif not has_chinese1 and has_chinese2:
                return cell_value
    
    # 如果不是上述格式，返回原内容
    return cell_value

def translate_worksheet(ws, app_id, app_key, columns, start_row=1, end_row=None):
    """翻译Excel工作表的指定列，英文在上中文在下"""
    # 确定结束行
    if end_row is None or end_row == 0:
        end_row = ws.max_row
    
    # 验证行范围
    if start_row < 1:
        start_row = 1
    if end_row > ws.max_row:
        end_row = ws.max_row
    if start_row > end_row:
        print_error(f"起始行({start_row})不能大于结束行({end_row})")
        return False, 0, 0
    
    # 样式设置
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    highlight_font = Font(color="FF0000", bold=True)
    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    
    total_translated = 0
    total_errors = 0
    total_skipped = 0
    start_time = time.time()
    
    # 记录需要调整列宽的列
    columns_to_adjust = {}
    
    # 遍历所有需要翻译的列
    for col_letter in columns:
        try:
            col_index = column_index_from_string(col_letter)
        except ValueError:
            print_error(f"无效的列标识符 '{col_letter}'，已跳过")
            continue
        
        # 初始化列宽调整数据
        columns_to_adjust[col_index] = {
            'max_length': 0,
            'col_letter': col_letter
        }
        
        # 统计该列需要翻译的单元格数量
        translatable_cells = 0
        for row_idx in range(start_row, end_row + 1):
            cell = ws.cell(row=row_idx, column=col_index)
            if cell.value and isinstance(cell.value, str) and cell.value.strip():
                if not is_already_translated(cell.value):
                    translatable_cells += 1
        
        if translatable_cells == 0:
            print_warning(f"列 {col_letter} 在行 {start_row}-{end_row} 范围内没有需要翻译的新内容")
            continue
        
        print_success(f"处理列 {col_letter} - 发现 {translatable_cells} 个需要翻译的单元格")
        
        processed = 0
        error_54003_occurred = False
        
        # 遍历该列指定行范围内的单元格
        for row_idx in range(start_row, end_row + 1):
            cell = ws.cell(row=row_idx, column=col_index)
            
            # 跳过空单元格或非文本单元格
            if not cell.value or not isinstance(cell.value, str):
                continue
                
            # 去除首尾空格
            cell_value = cell.value.strip()
            if not cell_value:
                continue
                
            # 检查是否已经翻译过
            if is_already_translated(cell_value):
                # 检查并调整中英文顺序
                adjusted_value = check_and_adjust_translation_order(cell_value)
                if adjusted_value != cell_value:
                    cell.value = adjusted_value
                    print_warning(f"已调整单元格 {col_letter}{row_idx} 的中英文顺序")
                total_skipped += 1
                if total_skipped % 20 == 0:
                    print_warning(f"已跳过 {total_skipped} 个已翻译单元格...")
                continue
            
            # 翻译中文内容
            english_text = baidu_translate(cell_value, app_id, app_key)
            
            # 处理频率限制错误
            if english_text == "ERROR_54003":
                error_54003_occurred = True
                print_warning("检测到频率限制错误，等待10秒后重试...")
                time.sleep(10)
                english_text = baidu_translate(cell_value, app_id, app_key)
                
                if english_text == "ERROR_54003":
                    print_error("重试后仍遇到频率限制，跳过当前单元格")
                    total_errors += 1
                    cell.fill = highlight_fill
                    cell.font = highlight_font
                    continue
            
            if english_text and english_text != "ERROR_54003":
                # 英文在上，中文在下
                cell.value = f"{english_text}\n{cell_value}"
                cell.alignment = wrap_alignment
                processed += 1
                total_translated += 1
                
                # 更新最大列宽
                text_length = max(len(cell_value), len(english_text))
                if text_length > columns_to_adjust[col_index]['max_length']:
                    columns_to_adjust[col_index]['max_length'] = text_length
                
                # 显示进度
                if processed % 5 == 0 or processed == translatable_cells:
                    elapsed = time.time() - start_time
                    remaining = translatable_cells - processed
                    avg_time = elapsed / processed if processed > 0 else 0
                    eta = (remaining * avg_time) / 60 if remaining > 0 else 0
                    
                    progress = f"进度: {processed}/{translatable_cells} | 耗时: {elapsed:.1f}秒"
                    if eta > 0:
                        progress += f" | 预计剩余: {eta:.1f}分钟"
                    print(progress)
            
            # 添加延迟避免QPS限制
            time.sleep(1.5)
    
    # 自动调整列宽
    print_success("正在调整列宽...")
    for col_index, data in columns_to_adjust.items():
        col_letter = get_column_letter(col_index)
        max_length = data['max_length']
        
        # 计算合适的列宽
        column_width = min(max_length * 1.2 + 5, 50)
        ws.column_dimensions[col_letter].width = column_width
        print_success(f"列 {col_letter} 宽度设置为 {column_width:.1f}")
    
    # 自动调整行高
    print_success("正在调整行高...")
    base_line_height = 15  # 基础行高
    for row_idx in range(start_row, end_row + 1):
        max_lines = 1
        for col_letter in columns:
            try:
                col_index = column_index_from_string(col_letter)
                cell = ws.cell(row=row_idx, column=col_index)
                if cell.value and isinstance(cell.value, str):
                    lines = cell.value.count('\n') + 1
                    if lines > max_lines:
                        max_lines = lines
            except ValueError:
                continue
        
        # 设置行高（基础高度 + 每行额外高度）
        height = base_line_height + (max_lines - 1) * 10
        ws.row_dimensions[row_idx].height = min(height, 100)
        if max_lines > 1:
            print_success(f"行 {row_idx} 高度设置为 {height:.1f} (包含 {max_lines} 行内容)")
    
    elapsed = time.time() - start_time
    print_success(f"翻译完成: 新增{total_translated}条 | 跳过{total_skipped}条 | 错误{total_errors}条")
    return True, total_translated, total_errors, elapsed

def translate_excel(input_file, output_file, app_id, app_key, columns, 
                   start_row=1, end_row=None, sheet_names=None):
    """翻译Excel文件中的指定工作表"""
    # 先进行API测试连接
    print_header("正在测试百度翻译API连接...")
    test_result = baidu_translate("测试", app_id, app_key)
    if test_result is None or test_result == "ERROR_54003":
        print_error("API连接测试失败，请检查配置")
        return False
    
    print_success("API测试通过")
    print_success(f"测试翻译: '测试' → '{test_result}'")
    
    # 加载工作簿
    try:
        wb = openpyxl.load_workbook(input_file)
        print_success(f"成功加载Excel文件: {input_file}")
    except FileNotFoundError:
        print_error(f"找不到输入文件: {input_file}")
        return False
    except Exception as e:
        print_error(f"加载Excel文件错误: {str(e)}")
        return False
    
    # 确定要处理的工作表
    all_sheet_names = wb.sheetnames
    if sheet_names is None:
        sheet_names = all_sheet_names
        print_success(f"将处理所有工作表: {', '.join(sheet_names)}")
    else:
        # 过滤不存在的工作表
        valid_sheets = [name for name in sheet_names if name in all_sheet_names]
        invalid_sheets = [name for name in sheet_names if name not in all_sheet_names]
        
        if invalid_sheets:
            print_warning(f"以下工作表不存在: {', '.join(invalid_sheets)}")
        
        if not valid_sheets:
            print_error("没有有效的工作表可供处理")
            return False
        
        sheet_names = valid_sheets
        print_success(f"将处理工作表: {', '.join(sheet_names)}")
    
    total_translated = 0
    total_errors = 0
    total_time = 0
    processed_sheets = 0
    
    # 处理每个工作表
    for sheet_name in sheet_names:
        print_header(f"处理工作表: {sheet_name}")
        ws = wb[sheet_name]
        
        success, translated, errors, elapsed = translate_worksheet(
            ws, app_id, app_key, columns, start_row, end_row
        )
        
        if success:
            total_translated += translated
            total_errors += errors
            total_time += elapsed
            processed_sheets += 1
            
            print_success(f"工作表 '{sheet_name}' 处理完成!")
            print_success(f"翻译: {translated} 条 | 错误: {errors} 条 | 耗时: {elapsed:.1f}秒")
        else:
            print_error(f"工作表 '{sheet_name}' 处理失败")
    
    # 保存结果
    if processed_sheets > 0:
        try:
            wb.save(output_file)
            print_success(f"结果已保存到: {output_file}")
            print_success(f"总计翻译: {total_translated} 条 | 总计错误: {total_errors} 条 | 总耗时: {total_time:.1f}秒")
            return True
        except PermissionError:
            print_error("保存文件失败: 文件可能被其他程序打开，请关闭后重试")
        except Exception as e:
            print_error(f"保存文件失败: {str(e)}")
    
    return False

def backup_file(file_path):
    """创建文件备份"""
    if not os.path.exists(file_path):
        return file_path
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base, ext = os.path.splitext(file_path)
    backup_path = f"{base}_backup_{timestamp}{ext}"
    
    try:
        import shutil
        shutil.copy2(file_path, backup_path)
        print_success(f"已创建备份文件: {backup_path}")
        return backup_path
    except Exception as e:
        print_error(f"创建备份失败: {str(e)}")
        return file_path

if __name__ == "__main__":
    # ===== 配置区域 =====
    INPUT_EXCEL = "Case Lesson Learn-fabbump20250527(1).xlsx"    # 输入文件名
    OUTPUT_EXCEL = "Case Lesson Learn-fab&bump-6月_translated.xlsx"  # 输出文件名
    
    # 要翻译的工作表名称（None表示所有工作表）
    SHEET_NAMES = ["RT301"]  # 可以指定多个工作表
    
    # 要翻译的列（可以是多个列，如 ['A', 'C', 'E']）
    COLUMNS_TO_TRANSLATE = ['D']  # 修改这里添加/更改要翻译的列
    
    # 翻译行范围设置
    START_ROW = 2     # 起始行（包含，从1开始计数）
    END_ROW = None     # 结束行（包含），设为None表示到末尾
    
    # 百度翻译API凭证
    BAIDU_APP_ID = "20250623002388339"
    BAIDU_APP_KEY = "60Nv9gYwakpQ3D16Mp3A"
    # ===================
    
    print_header("Excel翻译工具启动")
    print_success(f"当前时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    # 验证凭证是否已配置
    if BAIDU_APP_ID == "你的百度AppID" or BAIDU_APP_KEY == "你的百度AppKey":
        print_error("请先配置百度翻译API凭证!")
        print_warning("1. 访问 https://api.fanyi.baidu.com")
        print_warning("2. 注册/登录后创建应用")
        print_warning("3. 获取AppID和AppKey并填入代码")
        print_warning("4. 添加您的IP到白名单")
        sys.exit(1)
    
    # 创建备份
    print_header("创建文件备份")
    backup_path = backup_file(INPUT_EXCEL)
    
    # 执行翻译
    print_header("开始翻译处理")
    print_success(f"输入文件: {INPUT_EXCEL}")
    print_success(f"输出文件: {OUTPUT_EXCEL}")
    print_success(f"翻译列: {', '.join(COLUMNS_TO_TRANSLATE)}")
    print_success(f"起始行: {START_ROW}, 结束行: {'末尾' if END_ROW is None else END_ROW}")
    
    success = translate_excel(
        INPUT_EXCEL, 
        OUTPUT_EXCEL,
        BAIDU_APP_ID,
        BAIDU_APP_KEY,
        COLUMNS_TO_TRANSLATE,
        START_ROW,
        END_ROW,
        SHEET_NAMES
    )
    
    if success:
        print_header("翻译处理完成!")
        print_success(f"输出文件已保存: {OUTPUT_EXCEL}")
        print_success(f"原始文件备份: {backup_path}")
    else:
        print_header("翻译处理失败")
        print_error("请检查错误信息并修正后重试")
    
    # 在Windows系统上保持窗口打开
    if os.name == 'nt':
        input("\n按Enter键退出...")

    # 百度翻译API凭证
    #BAIDU_APP_ID = "20250623002388339"
    #BAIDU_APP_KEY = "60Nv9gYwakpQ3D16Mp3A"
    
    