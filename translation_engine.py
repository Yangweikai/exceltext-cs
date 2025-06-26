import re
import requests
import random
import hashlib
import openpyxl
import time
import os
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import datetime

class TranslationTask:
    def __init__(self, task_id, filename, columns, start_row, end_row, sheet_names, app_id, app_key):
        self.task_id = task_id
        self.filename = filename
        self.columns = columns
        self.start_row = start_row
        self.end_row = end_row
        self.sheet_names = sheet_names
        self.app_id = app_id
        self.app_key = app_key
        self.status = "pending"  # pending, running, completed, failed
        self.progress = 0
        self.total_cells = 0
        self.translated_cells = 0
        self.error_cells = 0
        self.skipped_cells = 0
        self.current_sheet = ""
        self.current_cell = ""
        self.message = ""
        self.output_file = ""
        self.start_time = None
        self.end_time = None

def baidu_translate(query, app_id, app_key):
    """使用百度翻译API进行中译英，带有频率控制和错误处理"""
    salt = random.randint(32768, 65536)
    sign_str = app_id + query + str(salt) + app_key
    sign = hashlib.md5(sign_str.encode()).hexdigest()
    
    url = "https://fanyi-api.baidu.com/api/trans/vip/translate"
    params = {
        'q': query,
        'from': 'zh',
        'to': 'en',
        'appid': app_id,
        'salt': salt,
        'sign': sign
    }
    
    try:
        response = requests.get(url, params=params, timeout=25)
        response.raise_for_status()
        result = response.json()
        
        if 'error_code' in result:
            error_code = result['error_code']
            if error_code == '54003':  # 访问频率受限
                return "ERROR_54003"
            elif error_code == '52003':  # 未授权用户
                return None
            else:
                return None
        
        return result['trans_result'][0]['dst']
    
    except requests.exceptions.RequestException as e:
        return None
    except ValueError:
        return None
    except Exception as e:
        return None

def is_already_translated(cell_value):
    """检查单元格是否已经包含翻译内容"""
    if not isinstance(cell_value, str):
        return False
    
    # 检查是否包含英文括号格式：中文(英文)
    if re.search(r'\([a-zA-Z].*\)$', cell_value):
        return True
    
    # 检查是否包含换行分隔格式：英文\n中文（检查第一部分是否有英文）
    if '\n' in cell_value:
        parts = cell_value.split('\n')
        if len(parts) >= 2:
            first_part = parts[0].strip()
            if any(c.isalpha() for c in first_part):  # 检查第一部分是否包含英文字母
                return True
    
    return False

def check_and_adjust_translation_order(cell_value):
    """检查并调整中英文翻译顺序，确保英文在中文上方"""
    if not isinstance(cell_value, str):
        return cell_value
    
    # 处理括号格式：中文(英文)
    bracket_match = re.match(r'^(.*?)\s*\(([A-Za-z\s].*?)\)\s*$', cell_value)
    if bracket_match:
        chinese_part = bracket_match.group(1).strip()
        english_part = bracket_match.group(2).strip()
        return f"{english_part}\n{chinese_part}"
    
    # 处理换行格式
    if '\n' in cell_value:
        parts = [part.strip() for part in cell_value.split('\n') if part.strip()]
        if len(parts) == 2:
            part1, part2 = parts
            # 检查哪部分是中文，哪部分是英文
            has_chinese1 = bool(re.search(r'[\u4e00-\u9fa5]', part1))
            has_chinese2 = bool(re.search(r'[\u4e00-\u9fa5]', part2))
            
            # 如果part1是中文，part2是英文，则交换
            if has_chinese1 and not has_chinese2:
                return f"{part2}\n{part1}"
            # 如果已经是英文在上，中文在下，则不改变
            elif not has_chinese1 and has_chinese2:
                return cell_value
    
    # 如果不是上述格式，返回原内容
    return cell_value

def translate_worksheet_with_progress(ws, task, columns, start_row=1, end_row=None):
    """翻译Excel工作表的指定列，带进度更新"""
    # 确定结束行
    if end_row is None or end_row == 0:
        end_row = ws.max_row
    
    # 验证行范围
    if start_row < 1:
        start_row = 1
    if end_row > ws.max_row:
        end_row = ws.max_row
    if start_row > end_row:
        task.message = f"起始行({start_row})不能大于结束行({end_row})"
        return False
    
    # 样式设置
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    highlight_font = Font(color="FF0000", bold=True)
    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    
    # 记录需要调整列宽的列
    columns_to_adjust = {}
    
    # 遍历所有需要翻译的列
    for col_letter in columns:
        try:
            col_index = column_index_from_string(col_letter)
        except ValueError:
            continue
        
        # 初始化列宽调整数据
        columns_to_adjust[col_index] = {
            'max_length': 0,
            'col_letter': col_letter
        }
        
        # 遍历该列指定行范围内的单元格
        for row_idx in range(start_row, end_row + 1):
            if task.status == "failed":
                return False
                
            cell = ws.cell(row=row_idx, column=col_index)
            
            # 跳过空单元格或非文本单元格
            if not cell.value or not isinstance(cell.value, str):
                continue
                
            # 去除首尾空格
            cell_value = cell.value.strip()
            if not cell_value:
                continue
            
            task.current_cell = f"{col_letter}{row_idx}"
            
            # 检查是否已经翻译过
            if is_already_translated(cell_value):
                # 检查并调整中英文顺序
                adjusted_value = check_and_adjust_translation_order(cell_value)
                if adjusted_value != cell_value:
                    cell.value = adjusted_value
                task.skipped_cells += 1
                # 更新进度
                total_processed = task.translated_cells + task.error_cells + task.skipped_cells
                if task.total_cells > 0:
                    old_progress = task.progress
                    task.progress = min(100, int((total_processed / task.total_cells) * 100))
                    
                    # 更新状态信息，显示更详细的进度
                    if task.translated_cells > 0 or task.skipped_cells > 0:
                        task.message = f"正在处理工作表: {task.current_sheet} | 当前单元格: {task.current_cell} | 已翻译: {task.translated_cells} | 已跳过: {task.skipped_cells} | 错误: {task.error_cells}"
                    
                    # 调试信息
                    if task.progress > 100:
                        print(f"警告: 进度超过100% - 总单元格: {task.total_cells}, 已处理: {total_processed}, 进度: {task.progress}%")
                    elif task.progress < old_progress:
                        print(f"警告: 进度回退 - 从 {old_progress}% 到 {task.progress}%")
                continue
            
            # 翻译中文内容
            english_text = baidu_translate(cell_value, task.app_id, task.app_key)
            
            # 处理频率限制错误
            if english_text == "ERROR_54003":
                task.message = "检测到频率限制错误，等待10秒后重试..."
                time.sleep(10)
                english_text = baidu_translate(cell_value, task.app_id, task.app_key)
                
                if english_text == "ERROR_54003":
                    task.error_cells += 1
                    cell.fill = highlight_fill
                    cell.font = highlight_font
                    # 更新进度
                    total_processed = task.translated_cells + task.error_cells + task.skipped_cells
                    if task.total_cells > 0:
                        old_progress = task.progress
                        task.progress = min(100, int((total_processed / task.total_cells) * 100))
                        
                        # 更新状态信息，显示更详细的进度
                        if task.translated_cells > 0 or task.skipped_cells > 0:
                            task.message = f"正在处理工作表: {task.current_sheet} | 当前单元格: {task.current_cell} | 已翻译: {task.translated_cells} | 已跳过: {task.skipped_cells} | 错误: {task.error_cells}"
                        
                        # 调试信息
                        if task.progress > 100:
                            print(f"警告: 进度超过100% - 总单元格: {task.total_cells}, 已处理: {total_processed}, 进度: {task.progress}%")
                        elif task.progress < old_progress:
                            print(f"警告: 进度回退 - 从 {old_progress}% 到 {task.progress}%")
                    continue
            
            if english_text and english_text != "ERROR_54003":
                # 英文在上，中文在下
                cell.value = f"{english_text}\n{cell_value}"
                cell.alignment = wrap_alignment
                task.translated_cells += 1
                
                # 更新最大列宽
                text_length = max(len(cell_value), len(english_text))
                if text_length > columns_to_adjust[col_index]['max_length']:
                    columns_to_adjust[col_index]['max_length'] = text_length
            
            # 更新进度
            total_processed = task.translated_cells + task.error_cells + task.skipped_cells
            if task.total_cells > 0:
                old_progress = task.progress
                task.progress = min(100, int((total_processed / task.total_cells) * 100))
                
                # 更新状态信息，显示更详细的进度
                if task.translated_cells > 0 or task.skipped_cells > 0:
                    task.message = f"正在处理工作表: {task.current_sheet} | 当前单元格: {task.current_cell} | 已翻译: {task.translated_cells} | 已跳过: {task.skipped_cells} | 错误: {task.error_cells}"
                
                # 调试信息
                if task.progress > 100:
                    print(f"警告: 进度超过100% - 总单元格: {task.total_cells}, 已处理: {total_processed}, 进度: {task.progress}%")
                elif task.progress < old_progress:
                    print(f"警告: 进度回退 - 从 {old_progress}% 到 {task.progress}%")
            
            # 添加延迟避免QPS限制
            time.sleep(1.5)
    
    # 自动调整列宽
    for col_index, data in columns_to_adjust.items():
        col_letter = get_column_letter(col_index)
        max_length = data['max_length']
        
        # 计算合适的列宽
        column_width = min(max_length * 1.2 + 5, 50)
        ws.column_dimensions[col_letter].width = column_width
    
    # 自动调整行高
    base_line_height = 15  # 基础行高
    for row_idx in range(start_row, end_row + 1):
        max_lines = 1
        for col_letter in columns:
            try:
                col_index = column_index_from_string(col_letter)
                cell = ws.cell(row=row_idx, column=col_index)
                if cell.value and isinstance(cell.value, str):
                    lines = cell.value.count('\n') + 1
                    if lines > max_lines:
                        max_lines = lines
            except ValueError:
                continue
        
        # 设置行高（基础高度 + 每行额外高度）
        height = base_line_height + (max_lines - 1) * 10
        ws.row_dimensions[row_idx].height = min(height, 100)
    
    return True

def translate_excel_with_progress(task):
    """翻译Excel文件中的指定工作表，带进度更新"""
    try:
        task.start_time = time.time()
        task.status = "running"
        
        # 先进行API测试连接
        task.message = "正在测试百度翻译API连接..."
        test_result = baidu_translate("测试", task.app_id, task.app_key)
        if test_result is None or test_result == "ERROR_54003":
            task.message = "API连接测试失败，请检查配置"
            task.status = "failed"
            return
        
        task.message = "API测试通过，开始翻译..."
        
        # 加载工作簿
        try:
            task.message = f"正在加载文件: {task.filename}"
            if not os.path.exists(task.filename):
                task.message = f"文件不存在: {task.filename}"
                task.status = "failed"
                return
                
            wb = openpyxl.load_workbook(task.filename)
            task.message = f"成功加载Excel文件"
        except FileNotFoundError:
            task.message = f"找不到输入文件: {task.filename}"
            task.status = "failed"
            return
        except Exception as e:
            task.message = f"加载Excel文件错误: {str(e)}"
            task.status = "failed"
            return
        
        # 确定要处理的工作表
        all_sheet_names = wb.sheetnames
        if task.sheet_names is None:
            task.sheet_names = all_sheet_names
        else:
            # 过滤不存在的工作表
            valid_sheets = [name for name in task.sheet_names if name in all_sheet_names]
            if not valid_sheets:
                task.message = "没有有效的工作表可供处理"
                task.status = "failed"
                return
            task.sheet_names = valid_sheets
        
        # 预计算所有工作表的总单元格数
        task.message = "正在计算需要翻译的单元格总数..."
        total_translatable = 0
        total_skippable = 0  # 统计已翻译的单元格数
        
        for sheet_name in task.sheet_names:
            ws = wb[sheet_name]
            for col_letter in task.columns:
                try:
                    col_index = column_index_from_string(col_letter)
                    for row_idx in range(task.start_row, (task.end_row or ws.max_row) + 1):
                        cell = ws.cell(row=row_idx, column=col_index)
                        if cell.value and isinstance(cell.value, str) and cell.value.strip():
                            if is_already_translated(cell.value):
                                total_skippable += 1
                            else:
                                total_translatable += 1
                except ValueError:
                    continue
        
        # 总处理单元格数 = 需要翻译的 + 需要跳过的
        total_cells_to_process = total_translatable + total_skippable
        
        task.total_cells = total_cells_to_process
        task.translated_cells = 0
        task.error_cells = 0
        task.skipped_cells = 0
        task.progress = 0
        
        if total_cells_to_process == 0:
            task.message = "没有需要处理的内容"
            task.status = "completed"
            task.end_time = time.time()
            return
        
        task.message = f"发现 {total_translatable} 个需要翻译的单元格，{total_skippable} 个已翻译的单元格"
        
        # 生成输出文件名
        base_name = os.path.splitext(os.path.basename(task.filename))[0]
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        task.output_file = os.path.join("uploads", f"{base_name}_translated_{timestamp}.xlsx")
        
        # 处理每个工作表
        for sheet_name in task.sheet_names:
            if task.status == "failed":
                break
                
            task.current_sheet = sheet_name
            task.message = f"正在处理工作表: {sheet_name}"
            
            ws = wb[sheet_name]
            
            success = translate_worksheet_with_progress(
                ws, task, task.columns, task.start_row, task.end_row
            )
            
            if not success:
                task.status = "failed"
                return
        
        # 保存结果
        if task.status != "failed":
            try:
                wb.save(task.output_file)
                task.message = "翻译完成，文件已保存"
                task.status = "completed"
                task.end_time = time.time()
                task.progress = 100  # 确保进度显示100%
            except Exception as e:
                task.message = f"保存文件失败: {str(e)}"
                task.status = "failed"
        
    except Exception as e:
        task.message = f"翻译过程中发生错误: {str(e)}"
        task.status = "failed" 