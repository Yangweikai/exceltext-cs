from flask import Flask, render_template, request, jsonify, send_file, Response
import os
import uuid
import threading
import time
from werkzeug.utils import secure_filename
from translation_engine import TranslationTask, translate_excel_with_progress

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-here'
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# 确保上传目录存在
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# 存储翻译任务状态
translation_tasks = {}

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/favicon.ico')
def favicon():
    # 返回一个简单的SVG图标，避免404错误
    svg_icon = '''<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 100 100">
        <rect width="100" height="100" fill="#667eea"/>
        <text x="50" y="65" font-size="50" text-anchor="middle" fill="white">📊</text>
    </svg>'''
    return Response(svg_icon, mimetype='image/svg+xml')

@app.route('/upload', methods=['POST'])
def upload_file():
    try:
        if 'file' not in request.files:
            return jsonify({'error': '没有选择文件'}), 400
        
        file = request.files['file']
        if file is None or file.filename == '':
            return jsonify({'error': '没有选择文件'}), 400
        
        if file.filename is None or not file.filename.endswith('.xlsx'):
            return jsonify({'error': '请选择Excel文件(.xlsx)'}), 400
        
        # 保存文件
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        # 读取Excel文件信息
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
            
            wb = openpyxl.load_workbook(filepath)
            sheet_names = wb.sheetnames
            
            # 获取第一个工作表的信息来估算列数
            ws = wb.active
            if ws is None:
                return jsonify({'error': '无法读取工作表'}), 400
                
            max_col = ws.max_column
            columns = [get_column_letter(i) for i in range(1, min(max_col + 1, 27))]  # A-Z
            
            return jsonify({
                'success': True,
                'filename': filename,
                'filepath': filepath,
                'sheet_names': sheet_names,
                'columns': columns,
                'max_row': ws.max_row
            })
        except Exception as e:
            return jsonify({'error': f'读取Excel文件失败: {str(e)}'}), 400
            
    except Exception as e:
        return jsonify({'error': f'上传失败: {str(e)}'}), 500

@app.route('/start_translation', methods=['POST'])
def start_translation():
    try:
        data = request.json
        if data is None:
            return jsonify({'error': '无效的请求数据'}), 400
        
        # 创建翻译任务
        task_id = str(uuid.uuid4())
        task = TranslationTask(
            task_id=task_id,
            filename=data.get('filepath', ''),
            columns=data.get('columns', []),
            start_row=int(data.get('start_row', 1)),
            end_row=int(data.get('end_row', 0)) if data.get('end_row') else None,
            sheet_names=data.get('sheet_names', []),
            app_id=data.get('app_id', ''),
            app_key=data.get('app_key', '')
        )
        
        translation_tasks[task_id] = task
        
        # 在后台线程中执行翻译
        thread = threading.Thread(target=translate_excel_with_progress, args=(task,))
        thread.daemon = True
        thread.start()
        
        return jsonify({'success': True, 'task_id': task_id})
        
    except Exception as e:
        return jsonify({'error': f'启动翻译失败: {str(e)}'}), 500

@app.route('/progress/<task_id>')
def get_progress(task_id):
    if task_id not in translation_tasks:
        return jsonify({'error': '任务不存在'}), 404
    
    task = translation_tasks[task_id]
    
    elapsed_time = 0
    if task.start_time:
        if task.end_time:
            elapsed_time = task.end_time - task.start_time
        else:
            elapsed_time = time.time() - task.start_time
    
    return jsonify({
        'status': task.status,
        'progress': task.progress,
        'total_cells': task.total_cells,
        'translated_cells': task.translated_cells,
        'error_cells': task.error_cells,
        'skipped_cells': task.skipped_cells,
        'current_sheet': task.current_sheet,
        'current_cell': task.current_cell,
        'message': task.message,
        'elapsed_time': round(elapsed_time, 1),
        'output_file': task.output_file if task.status == "completed" else None
    })

@app.route('/download/<task_id>')
def download_file(task_id):
    if task_id not in translation_tasks:
        return jsonify({'error': '任务不存在'}), 404
    
    task = translation_tasks[task_id]
    
    if task.status != "completed" or not task.output_file:
        return jsonify({'error': '文件未准备好'}), 400
    
    if not os.path.exists(task.output_file):
        return jsonify({'error': '文件不存在'}), 404
    
    return send_file(task.output_file, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000) 